from aiogram import Router
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile

from run import bot, db
from config import ADMIN

from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

export_router = Router()


@export_router.message(Command("export_all"))
async def search(message: Message):
    admin_id = str(message.from_user.id)
    if admin_id in ADMIN:
        result = await db.select_all_users()
        if result:
            columns, data = result  # Unpack the returned tuple

            # Convert the data to DataFrame with column names
            excel_file = pd.DataFrame(data, columns=columns)

            # Create a BytesIO object to store the Excel file in memory
            output = BytesIO()

            # Use the ExcelWriter context manager to write the DataFrame to Excel
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                excel_file.to_excel(writer, sheet_name='Users', index=False)

                # Access the worksheet to apply styling
                worksheet = writer.sheets['Users']

                # Center alignment style
                center_aligned = Alignment(horizontal='center', vertical='center')

                # Style for headers (centered and bold)
                header_font = Font(bold=True, size=12)  # Added size for better readability

                # Apply styling to all cells including headers
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = center_aligned

                # Apply additional header styling
                for cell in worksheet[1]:
                    cell.font = header_font

                # Adjust column widths based on content
                for idx, col in enumerate(excel_file.columns):
                    # Get the maximum length of the column name and its content
                    max_length = max(
                        excel_file[col].astype(str).apply(len).max(),  # max content length
                        len(str(col))  # length of column name
                    )
                    # Add extra space
                    adjusted_width = max_length + 3  # Increased padding for better readability
                    # Set the column width
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width

                # Freeze the header row
                worksheet.freeze_panes = 'A2'

            # Get the bytes from the BytesIO object
            excel_bytes = output.getvalue()

            # Create a BufferedInputFile
            document = BufferedInputFile(excel_bytes, filename="UserInfo.xlsx")

            # Send the document
            await bot.send_document(
                chat_id=admin_id,
                document=document,
                caption="Here's your All User Info."
            )
        else:
            await bot.send_message(chat_id=admin_id, text="No user data found.")
    else:
        await message.answer("You don't have permission")